import pandas as pd
import re
from datetime import timedelta, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Color, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

def allocateTasksToWeeks(tasks):
    colWeekHours = [40] #first column, representing a work week, or 40 hours
    chronogram = [] 

    for task in tasks:
        weeks = len(colWeekHours)
        taskRow = ['_'] * weeks #current row times the weeks needed 
        while task > 0: #While the task has hours left to assing
            for i in range(len(colWeekHours)):
                if task <= colWeekHours[i]: #task needs less hours than available in current work week
                    colWeekHours[i] -= task
                    taskRow[i] = 'X'
                    task = 0 #Task hours fully allocated
                    break
                else:        #Task needs more hours than available in current work week
                    if colWeekHours[i] > 0:
                        task -= colWeekHours[i]
                        taskRow[i] = 'X'
                        colWeekHours[i] = 0

            #If task still has hours left not allocated, add new week
            if task > 0:   
                colWeekHours.append(40)
                taskRow.append('_')    #Extend task row for the new week

        #Update chronogram with next task
        chronogram.append(taskRow) 

         #Add weeks not used by tasks 
        for row in chronogram: 
            while len(row) < len(taskRow):
                row.append('_')

    return chronogram

# Function to validate the date format MM/DD
def validate_date(date_text):
    try:
        datetime.strptime(date_text, '%m/%d')
        return True
    except ValueError:
        return False
    
# Function to calculate date ranges for each week, spanning 7 days each
def get_week_dates(start_date, num_weeks, year):
    if not start_date:  # If start_week is empty
        # Return week numbers instead of dates
        return [f'Week {i+1}' for i in range(num_weeks + 2)]
    else:
        week_dates = []
        start_date = datetime.strptime(f"{start_date}/{year}", "%m/%d/%Y")
        for i in range(num_weeks + 2):
            end_date = start_date + timedelta(days=6)
            week_dates.append(f'{start_date.strftime("%d/%b")} - {end_date.strftime("%d/%b")}')
            start_date = end_date + timedelta(days=1)
        return week_dates

def add_task_dates(chronogram, start_date, ws, year):
    if not start_date:
        return  # If no start_date is provided, we cannot calculate task dates.

    # Convert start_date string to a datetime object, including the year.
    start_date_with_year = f"{start_date}/{year}"
    current_start_date = datetime.strptime(start_date_with_year, "%m/%d/%Y")

    # Get week dates to match with the 'X' marks in the chronogram.
    week_dates = get_week_dates(start_date, len(chronogram[0]), year)

    for index, task_row in enumerate(chronogram, start=1):
        # Find the indices of the 'X's in the task row to determine start and end dates.
        x_indices = [i for i, x in enumerate(task_row) if x == 'X']
        if x_indices:
            # Get the week date range for the first and last 'X'.
            start_week_range = week_dates[x_indices[0]].split(' - ')[0]
            end_week_range = week_dates[x_indices[-1]].split(' - ')[1]

            # Convert week range strings to datetime objects.
            task_start_date = datetime.strptime(f"{start_week_range}/{year}", "%d/%b/%Y")
            task_end_date = datetime.strptime(f"{end_week_range}/{year}", "%d/%b/%Y")

            # Use the actual start_date as the task's start date if it's the first task.
            if index == 1:
                task_start_date = max(task_start_date, current_start_date)

            # Set the start and end date cells.
            ws.cell(row=index+3, column=4, value=task_start_date.strftime("%m/%d"))
            ws.cell(row=index+3, column=5, value=task_end_date.strftime("%m/%d"))

# Function to adjust column widths and text wrapping in your Excel file
def adjust_column_settings(ws):
    # Set wider column widths for specific columns
    column_widths = {
        'B': 5,  # Tasks column
        'C': 30,  # Activity column
        'D': 10,  # Start Date column
        'E': 10,  # End Date column
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Enable text wrapping
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=2, max_col=3):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

def chronogramToExcel(chronogram, year, start_week, activity_names, filename="chronogram.xlsx"):
    # Start from column F (which is index 5 in zero-indexed systems)
    start_col_index = 6

    # Create DataFrame from chronogram
    df = pd.DataFrame(chronogram)

    # Insert empty columns at the beginning to shift the data to start from column F
    for col in range(start_col_index - 1):  # -1 because the DataFrame already starts with index 1
        df.insert(col, 'Empty{}'.format(col), [''] * df.shape[0])

    # Write DataFrame to Excel file without the index and header
    df.to_excel(filename, index=False, header=False)

    # If year is not provided, use the current year
    if not year:
        year = datetime.now().year

    # Open Excel file and color cells
    wb = Workbook()
    ws = wb.active  # Get the active sheet

    # Merge cells for the year header starting from column F
    last_data_column = len(df.columns) + 1
    ws.merge_cells(start_row=1, start_column=start_col_index, end_row=1, end_column=last_data_column)

    # Set the value for the year header and apply styles
    year_cell = ws.cell(row=1, column=start_col_index)
    year_cell.value = str(year)
    year_cell.alignment = Alignment(horizontal='center', vertical='center')
    year_cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    year_cell.font = Font(color="FFFFFF", bold=True)

    # Insert month headers and week date ranges
    week_dates = get_week_dates(start_week, len(df.columns) - start_col_index, year)
    row_offset = 2  # This is where the headers will start in the Excel sheet
    months = {}
    # Merge cells for the "Task" header from B1 to B3
    ws.merge_cells(start_row=1, start_column=2, end_row=3, end_column=2)
    # Apply styles and set the value for the "Task" header in the top-left cell of the merged area
    task_header_cell = ws.cell(row=1, column=2)
    task_header_cell.alignment = Alignment(horizontal='center', vertical='bottom')
    task_header_cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    task_header_cell.font = Font(color="FFFFFF", bold=True)
    task_header_cell.value = "Tasks"  # Even though this sets B1, it visually appears in B3 due to the merged cells

    # Add the task numbers in column B, starting from the row where the yellow 'X' begins
    row_offset_for_tasks = 4  # assuming the yellow 'X' begins at row 4
    for task_num, task_row in enumerate(chronogram, start=1):
        task_start_row = row_offset_for_tasks + task_num - 1
        task_cell = ws.cell(row=task_start_row, column=2)  # Start from the adjusted offset row
        #task_cell.value = f'Task {task_num}'
        task_cell.value = task_num

    # Determine the actual number of weeks with tasks (plus the extra week)
    actual_weeks_with_tasks = len(df.columns) - start_col_index + 1
    
    # Assign month headers and week headers starting from column F
    for i, date_range in enumerate(week_dates, start=start_col_index):
        if start_week:
            month_name = datetime.strptime(date_range.split(' - ')[0], "%d/%b").strftime("%B")
            if month_name not in months:
                months[month_name] = {'start': i, 'end': i}
            else:
                months[month_name]['end'] = i
        else:
            # No starting week provided, so label the "months" as Month 1, Month 2, etc.
            # Month is determined by the week number (i - start_col_index + 1)
            # Assumes each month has 4 weeks
            month_num = ((i - start_col_index) // 4) + 1
            month_name = f'Month {month_num}'
            if month_name not in months:
                # Start a new month
                months[month_name] = {'start': i, 'end': i}
            # Ensure that the month does not extend beyond the actual weeks with tasks
            if i < actual_weeks_with_tasks:
                months[month_name]['end'] = i
            # Set the week header as Week 1, Week 2, etc.
            week_cell = ws.cell(row=row_offset + 1, column=i)
            week_cell.value = f'Week {i - start_col_index + 1}'

        # Set week headers
        week_cell = ws.cell(row=row_offset + 1, column=i)
        week_cell.value = date_range
        week_cell.alignment = Alignment(horizontal='center')
        week_cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
        week_cell.font = Font(color="FFFFFF")

    # Now, outside the loop, adjust month end values, merge cells, and fill for "monthly" headers if no start_week provided
    if not start_week:
        # First, adjust the end of each month to match the last week with a task or the end of the 4-week block
        for month_name, month_range in months.items():
            month_end_week = month_range['start'] + 3  # Default month span of 4 weeks
            if month_end_week > actual_weeks_with_tasks + start_col_index - 1:
                month_end_week = actual_weeks_with_tasks + start_col_index  # Adjust to not go past the actual weeks with tasks
            months[month_name]['end'] = month_end_week

        # Merge and fill cells for each month
        for month_name, month_range in months.items():
            ws.merge_cells(start_row=row_offset, start_column=month_range['start'], end_row=row_offset, end_column=month_range['end'])
            for col_index in range(month_range['start'], month_range['end'] + 1):
                month_cell = ws.cell(row=row_offset, column=col_index)
                month_cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
                month_cell.font = Font(color="FFFFFF")
                if col_index == month_range['start']:  # Only write the month name in the first cell
                    month_cell.value = month_name
                    month_cell.alignment = Alignment(horizontal='center')

    # Merge cells for month headers
    for month, cols in sorted(months.items()):
        ws.merge_cells(start_row=row_offset, start_column=cols['start'], end_row=row_offset, end_column=cols['end'])
        month_cell = ws.cell(row=row_offset, column=cols['start'])
        month_cell.value = month
        month_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        month_cell.fill = PatternFill(start_color="0070c0", end_color="0070c0", fill_type="solid")
        month_cell.font = Font(color="FFFFFF")

    # Adjust row_offset for tasks below the week headers
    row_offset += 2

    # Add tasks to the Excel sheet
    for index, row in enumerate(chronogram, start=row_offset):
        for col_index, value in enumerate(row, start=start_col_index):
            task_cell = ws.cell(row=index, column=col_index)
            if value == 'X':
                task_cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

    # Set column widths for the data starting from column F
    column_width = 18
    for col in ws.iter_cols(min_col=start_col_index, max_col=ws.max_column, min_row=1, max_row=ws.max_row):
        for cell in col:
            ws.column_dimensions[get_column_letter(cell.column)].width = column_width
    
    # Create and style "Activity" header in column C
    ws.merge_cells(start_row=1, start_column=3, end_row=3, end_column=3)
    activity_header_cell = ws.cell(row=1, column=3)
    activity_header_cell.value = "Activity"
    activity_header_cell.alignment = Alignment(horizontal='center', vertical='bottom')
    activity_header_cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    activity_header_cell.font = Font(color="FFFFFF", bold=True)
    
    if activity_names:
        # Add the activity names to column C, starting from the 4th row to match the task rows
        for index, activity_name in enumerate(activity_names, start=4):
            ws.cell(row=index, column=3, value=activity_name)

    # Create and style "Start Date" and "End Date" headers
    ws.merge_cells(start_row=1, start_column=4, end_row=3, end_column=4)  # Merge cells for "Start Date"
    ws.merge_cells(start_row=1, start_column=5, end_row=3, end_column=5)  # Merge cells for "End Date"
        
    start_date_header_cell = ws.cell(row=1, column=4)
    end_date_header_cell = ws.cell(row=1, column=5)
    
    start_date_header_cell.value = "Start Date"
    end_date_header_cell.value = "End Date"
    
    for cell in [start_date_header_cell, end_date_header_cell]:
        cell.alignment = Alignment(horizontal='center', vertical='bottom')
        cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)

    # Before saving the workbook, call the function to add task start dates
    add_task_dates(chronogram, start_week, ws, year)  # Pass 'year' as well

    # Adjust column settings before saving
    adjust_column_settings(ws)

    # Save the workbook
    wb.save(filename)
    df.to_csv("chronogram.csv", index=False)  # Also save as CSV

# Ask user for the year for the Gantt Chart
yearInput = input("Add the year for the Gantt Chart (leave empty if using current year): ").strip()
year = int(yearInput) if yearInput else datetime.now().year

# Prompt the user for the starting week, now expecting MM/DD format
start_week = input("Add the starting week (MM/DD) (leave empty if not): ").strip()
while start_week and not validate_date(start_week):
    start_week = input("The format is incorrect. Please use MM/DD format or leave empty: ").strip()

# Ask user for input (hours as separated values by comma)
taskHoursInput = input("Add tasks hours (as comma-separated values): ")
tasks = [int(x.strip()) for x in re.split(r'[,\s]+', taskHoursInput) if x.strip()]

# Ask user for input (activity names as comma-separated values, or leave empty)
activityNamesInput = input("Add the activities (as comma-separated values, or leave empty): ")
activity_names = [x.strip() for x in re.split(r',\s*|\,', activityNamesInput) if x.strip()]

# Generate the chronogram from user input
chronogram = allocateTasksToWeeks(tasks)

# Call the function to save the chronogram to an Excel file
chronogramToExcel(chronogram, year, start_week if start_week.strip() else "", activity_names, "chronogram.xlsx")





        


